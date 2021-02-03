from openpyxl import *
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
top=Tk()
top.geometry("1000x750")
top.title("Student Registration Form")
wb = load_workbook('simi.xlsx')
sheet = wb.active

def onemore():
    root = Tk()
    root.title("Upload documents")
    root.geometry("800x400")
    top.destroy()
    def browsefunc1():
        filename1 = filedialog.askopenfilename()
        pathlabel1.config(text=filename1)
    def browsefunc2():
        filename1 = filedialog.askopenfilename()
        pathlabel2.config(text=filename1)
    def browsefunc3():
        filename1 = filedialog.askopenfilename()
        pathlabel3.config(text=filename1)
    def browsefunc5():
        msg1=messagebox.askyesno("Recheck!!","Are you sure to submit your details? \n No further changes can be done if you submit\n it. To recheck press NO.")
        if msg1:
            msg2=messagebox.showinfo("Congratulations!!","Congratulations, you have successfully submitted your registration form .")
            if msg2:
                root.destroy()
                
        else:
            pass
    def browsefunc4():
        msg2=messagebox.showinfo(".........","Your details are successfully saved.")  
        
                                 
    
    l1=Label(root,text="  UPLOAD YOUR DOCUMENTS  ")
    l1.grid(row=0,column=2,pady=25,padx=10)
    l1=Label(root,text="   1. Upload your Photograph : ")
    l1.grid(row=1,column=1,pady=15)
    b1 = Button(root, text="Browse", command=browsefunc1,bg="white")
    b1.grid(row=2,column=1)
    pathlabel1 = Label(root)
    pathlabel1.grid(row=2,column=2)


    l2=Label(root,text="2. Upload your Signature : ")
    l2.grid(row=4,column=1,pady=15)
    b2 = Button(root, text="Browse", command=browsefunc2,bg="white")
    b2.grid(row=5,column=1)
    pathlabel2 = Label(root)
    pathlabel2.grid(row=5,column=2)

    l3=Label(root,text="   3. Upload your Adhaar card : ")
    l3.grid(row=7,column=1,pady=15)
    b3 = Button(root, text="Browse", command=browsefunc3,bg="white")
    b3.grid(row=8,column=1)
    pathlabel3 = Label(root)
    pathlabel3.grid(row=8,column=2)

    l4=Label(root,text="    ")
    l4.grid(row=9,column=1,pady=15)
    b4=Button(root,text= "SAVE",bg="orange",width=10,command=browsefunc4)
    b4.grid(row=10,column=3)
    l4=Label(root,text="  ")
    l4.grid(row=10,column=4,pady=20)
    b4=Button(root,text= "SUBMIT",bg="green",width=10,command=browsefunc5)
    b4.grid(row=10,column=5)
    root.mainloop()


def excel(): 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 7
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 4
    sheet.column_dimensions['K'].width = 10
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['M'].width = 80
    sheet.column_dimensions['N'].width = 20
    sheet.column_dimensions['O'].width = 20
    sheet.column_dimensions['P'].width = 20
    sheet.column_dimensions['Q'].width = 10
    sheet.column_dimensions['R'].width = 12
    sheet.column_dimensions['S'].width = 20
    sheet.column_dimensions['T'].width = 40
    sheet.column_dimensions['U'].width = 20
    sheet.column_dimensions['V'].width = 20
    sheet.column_dimensions['W'].width = 20
    sheet.column_dimensions['X'].width = 20
    sheet.column_dimensions['Y'].width = 20
    sheet.column_dimensions['Z'].width = 20
  
    sheet.cell(row=1, column=1).value = "Name of the Applicant"
    sheet.cell(row=1, column=2).value = "Father's Name"
    sheet.cell(row=1, column=3).value = "Mother's Name"
    sheet.cell(row=1, column=4).value = "Father's Occupation"
    sheet.cell(row=1, column=5).value = "Mother's Occupation"
    sheet.cell(row=1, column=6).value = "Date Of Birth"
    sheet.cell(row=1, column=7).value = "Gender"
    sheet.cell(row=1, column=8).value = "Adhaar card no."
    sheet.cell(row=1, column=9).value = "Language Spoken"
    sheet.cell(row=1, column=10).value = "Age"
    sheet.cell(row=1, column=11).value = "Status"
    sheet.cell(row=1, column=12).value = "Category"
    sheet.cell(row=1, column=13).value = "Address"
    sheet.cell(row=1, column=14).value = "State"
    sheet.cell(row=1, column=15).value = "District"
    sheet.cell(row=1, column=16).value = "City"
    sheet.cell(row=1, column=17).value = "Pincode"
    sheet.cell(row=1, column=18).value = "Contact no."
    sheet.cell(row=1, column=19).value = "Another Contact no."
    sheet.cell(row=1, column=20).value = "Email"
    sheet.cell(row=1, column=21).value = "Last Class Passed"
    sheet.cell(row=1, column=22).value = "Board / University"
    sheet.cell(row=1, column=23).value = "Marks Obtained"
    sheet.cell(row=1, column=24).value = "Total Marks"
    sheet.cell(row=1, column=25).value = "Last Year Percentage"
    sheet.cell(row=1, column=26).value = "Stream Applied"
def ok7():
    str1=""
    if (var32.get()==1):
        str1=str1+"CSE"
    elif(var32.get()==2):
        str1=str1+"ECE"
    elif(var32.get()==3):
        str1=str1+"CIVIL"
    elif(var32.get()==4):
        str1=str1+"BIOTECH"
    else:
        str1=str1+"MECHANICAL"
    return str1

def ok6():
    str1=""
    if (var13.get()==1):
        str1=str1+"GENERAL"
    elif(var13.get()==2):
        str1=str1+"BCA"
    elif(var13.get()==3):
        str1=str1+"BCB"
    else:
        str1=str1+"SC / ST"
    return str1

def ok5():
    str1=""
    if (var12.get()==1):
        str1=str1+"10th"
    elif(var12.get()==2):
        str1=str1+"12th"
    elif(var12.get()==5):
        str1=str1+"Graduation"
    elif(var12.get()==6):
        str1=str1+"Post graduation"
    elif(var12.get()==3):
        str1=str1+"ITI"
    else:
        str1=str1+"Diploma"
    return str1

def ok4():
    str1=""
    if (var17.get()==1):
        str1=str1+"CBSE"
    elif(var17.get()==2):
        str1=str1+"ICSE"
    elif(var17.get()==3):
        str1=str1+"State govt"
    elif(var17.get()==4):
        str1=str1+"UNIVERSITY"   
    else:
        str1=str1+"UNIVERSITY AFF. COLLEGE"
    return str1
def ok3():
    str1=""
    if (var11.get()==1):
        str1=str1+"Male"
    elif(var11.get()==2):
        str1=str1+"Female"
    else:
        str1=str1+"Trans"
    return str1
def ok():
    str1=""
    if(var14.get())==1:
        str1=str1+" English"
    if(var15.get())==1:
        str1=str1+" Hindi"
    if(var16.get())==1:
        str1=str1+" Punjabi"
    return str1

def ok1():
    str2=""
    if(var27.get())==1:
        str2=str2+ "Married"
    if(var22.get())==1:
        str2=str2+ "Unmarried"
    return str2

def ok2():
    x=int(var24.get())
    y=int(var25.get())
    z=((x/y)*100)
    return str(z)+"%"

def insert():     
    current_row = sheet.max_row 
    current_column = sheet.max_column 
     
    sheet.cell(row=current_row + 1, column=1).value = (str(var1.get()) +" "+ str(var2.get()) + " " +str(var3.get())) 
    sheet.cell(row=current_row + 1, column=2).value = str(var4.get()) 
    sheet.cell(row=current_row + 1, column=3).value = str(var5.get()) 
    sheet.cell(row=current_row + 1, column=4).value = str(var6.get()) 
    sheet.cell(row=current_row + 1, column=5).value = str(var7.get()) 
    sheet.cell(row=current_row + 1, column=6).value = (str(var8.get())+"/"+str(var9.get())+"/"+str(var10.get())) 
    sheet.cell(row=current_row + 1, column=7).value = ok3()
    sheet.cell(row=current_row + 1, column=8).value = str(e18.get())
    sheet.cell(row=current_row + 1, column=9).value = ok()
    sheet.cell(row=current_row + 1, column=10).value = str(var26.get())
    sheet.cell(row=current_row + 1, column=11).value = ok1()
    sheet.cell(row=current_row + 1, column=12).value = ok6()
    sheet.cell(row=current_row + 1, column=13).value = str(t1.get("1.0",END))
    sheet.cell(row=current_row + 1, column=14).value = str(var18.get())
    sheet.cell(row=current_row + 1, column=15).value = str(var19.get())
    sheet.cell(row=current_row + 1, column=16).value = str(var20.get())
    sheet.cell(row=current_row + 1, column=17).value = str(var21.get())
    sheet.cell(row=current_row + 1, column=18).value = str(var28.get())
    sheet.cell(row=current_row + 1, column=19).value = str(var29.get())
    sheet.cell(row=current_row + 1, column=20).value = str(var30.get())
    sheet.cell(row=current_row + 1, column=21).value = ok4()
    sheet.cell(row=current_row + 1, column=22).value = ok5()
    sheet.cell(row=current_row + 1, column=23).value = str(var24.get())
    sheet.cell(row=current_row + 1, column=24).value = str(var25.get())
    sheet.cell(row=current_row + 1, column=25).value = ok2()
    sheet.cell(row=current_row + 1, column=26).value = ok7()
        
    wb.save('simi.xlsx') 
  
def percentage():
    x=int(var24.get())
    y=int(var25.get())
    z=((x/y)*100)
    print("Last Year Percentage : "+str(z)+"%")
def uff():
    print("PERSONAL DETAILS : \n")
    print("Name Of The Applicant : "+ str(var1.get()) +" "+ str(var2.get()) + " " +str(var3.get()))
    print("Father's Name : "+str(var4.get())+"\t\t\t"+ "Mother's Name : " +str(var5.get()))
    print("Father's Occupation : "+str(var6.get())+"\t\t"+ "Mother's Occupation : " +str(var7.get()))
    print("Date Of Birth : "+str(var8.get())+"/"+str(var9.get())+"/"+str(var10.get()))
    if (var11.get()==1):
        print("Gender : Male")
    elif(var11.get()==2):
        print("Gender : Female")
    else:
        print("Gender : Trans")
    str1 ="Language Spoken : "
    if(var14.get())==1:
        str1=str1+" English"
    if(var15.get())==1:
        str1=str1+" Hindi"
    if(var16.get())==1:
        str1=str1+" Punjabi"
    print(str1)
    print("Age : "+str(var26.get()))
    if(var27.get())==1:
         print("Status : Married")
    if(var22.get())==1:
        print("Status : Unmarried")
    print("")
    print("ADDRESS DETAILS : \n")
    print("Address : "+str(t1.get("1.0",END)))
    print("State : "+str(var18.get())+"\t\t\t\t\t\t"+"District : "+str(var19.get())+"\n"+"City : "+str(var20.get())+"\t\t\t\t\t\t"+"Pincode : "+str(var21.get()))
    print("\nCONTACT  DETAILS : \n")
    print("Contact no. : "+str(var28.get())+"\t\t\t\t"+"Another Contact no. : "+str(var29.get()))
    print("Email : "+str(var30.get()))
    print(" \nMARKS    DETAILS : \n")
    print("Marks Obtained : "+str(var24.get())+"\t\t\t\t\t\t"+"Total Marks : "+str(var25.get()))
    percentage()
    
var1 = StringVar()
var2 = StringVar()
var3 = StringVar()
var4 = StringVar()
var5 = StringVar()
var6 = StringVar()
var7 = StringVar()
var8 = IntVar()
var9 = IntVar()
var10 = IntVar()
var11 = IntVar()
var12 = IntVar()
var13 = IntVar()
var14 = IntVar()
var15 = IntVar()
var16 = IntVar()
var17 = IntVar()
var18 = StringVar()
var19 = StringVar()
var20 = StringVar()
var21 = StringVar()
var22 = IntVar()
var24 = StringVar()
var25 = StringVar()
var26 = IntVar()
var27 = IntVar()
var28 = StringVar()
var29 = StringVar()
var30 = StringVar()
var31 = IntVar()
var32 = IntVar()


def button1():
    msg1=messagebox.askyesno("Surity","Are you sure to move next ??\n If you want to go back you can press NO. ")
    if msg1:
        if (var31.get())!= 1:
            messagebox.showinfo("Warning","Please fill all the details carefully")
        else:
            messagebox.showinfo("Successful","You have successfully registered yourself")
            insert()
            onemore()
    else:
        pass
    
def button2():
    msg2=messagebox.askyesno("Warning", "Are you sure to cancel?")
    if msg2:
        top.destroy()
    else:
        pass

def button3():
    msg3=messagebox.askyesno("Print","Are you sure to print your details?")
    if msg3:
        uff()
    else:
        pass

l1=Label(top,text=" STUDENT   REGISTRATION   FORM ",bg="white",pady=10)
l1.grid(row=1,column=7)
l2=Label(top,text="PERSONAL DETAILS :",pady=10)
l2.grid(row=3,column=1)

l3=Label(top,text="First name :")
l3.grid(row=4,column=1)
e1=Entry(top,bd=1,textvariable=var1)
e1.grid(row=4,column=4)
l4=Label(top,text="Middle name :")
l4.grid(row=4,column=7)
e2=Entry(top,bd=1,textvariable=var2)
e2.grid(row=4,column=8)
l5=Label(top,text="Last name :")
l5.grid(row=4,column=9)
e3=Entry(top,bd=1,textvariable=var3)
e3.grid(row=4,column=10)

l6=Label(top,text="Father's name :")
l6.grid(row=6,column=1)
e4=Entry(top,bd=1,textvariable=var4)
e4.grid(row=6,column=4)
l7=Label(top,text="Mother's name :")
l7.grid(row=6,column=7)
e5=Entry(top,bd=1,textvariable=var5)
e5.grid(row=6,column=8)

l8=Label(top,text="Father's occupation :")
l8.grid(row=8,column=1)
e6=Entry(top,bd=1,textvariable=var6)
e6.grid(row=8,column=4)
l9=Label(top,text="Mother's occupation :")
l9.grid(row=8,column=7)
e7=Entry(top,bd=1,textvariable=var7)
e7.grid(row=8,column=8)

l10=Label(top,text="Date of birth :")
l10.grid(row=10,column=1)
l10=Label(top,text="DD")
l10.grid(row=10,column=2)
s1=Spinbox(top,from_=0,to=31,width=4,textvariable=var8)
s1.grid(row=10,column=3)
l10=Label(top,text="MM")
l10.grid(row=11,column=2)
s2=Spinbox(top,from_=0,to=12,width=4,textvariable=var9)
s2.grid(row=11,column=3)
l10=Label(top,text="YY")
l10.grid(row=12,column=2)
s3=Spinbox(top,from_=1980,to=2000,width=4,textvariable=var10)
s3.grid(row=12,column=3)

l11=Label(top,text="Gender :")
l11.grid(row=10,column=7)
r1=Radiobutton(top,text="Male  ",value=1,variable=var11)
r1.grid(row=10,column=8)
r1=Radiobutton(top,text="Female",value=2,variable=var11)
r1.grid(row=10,column=9)
r1=Radiobutton(top,text="Trans ",value=3,variable=var11)
r1.grid(row=10,column=10)

l12=Label(top,text="Language spoken :")
l12.grid(row=12,column=7)
c1=Checkbutton(top,text="English",onvalue=1,offvalue=0,variable=var14)
c1.grid(row=12,column=8)
c2=Checkbutton(top,text="Hindi",onvalue=1,offvalue=0,variable=var15)
c2.grid(row=12,column=9)
c3=Checkbutton(top,text="Punjabi",onvalue=1,offvalue=0,variable=var16)
c3.grid(row=12,column=10)

l13=Label(top,text="Adhaar card no. :")
l13.grid(row=13,column=1)
e18=Entry(top,bd=1)
e18.grid(row=13,column=4)


l19=Label(top,text=" ADDRESS    DETAILS : ",pady=10)
l19.grid(row=16,column=1)
l14=Label(top,text="Address :")
l14.grid(row=16,column=8)
t1=Text(top,width=20,height=4)
t1.grid(row=16,column=9)
l15=Label(top,text="State :")
l15.grid(row=17,column=1)
e8=Entry(top,bd=1,textvariable=var18)
e8.grid(row=17,column=4)
l16=Label(top,text="District :")
l16.grid(row=17,column=9)
e9=Entry(top,bd=1,textvariable=var19)
e9.grid(row=17,column=10)
l17=Label(top,text="City :")
l17.grid(row=18,column=1)
e10=Entry(top,bd=1,textvariable=var20)
e10.grid(row=18,column=4)
l18=Label(top,text="Pincode :")
l18.grid(row=18,column=9)
e11=Entry(top,bd=1,textvariable=var21)
e11.grid(row=18,column=10)

l20=Label(top,text=" EDUCATIONAL    DETAILS : ",pady=10)
l20.grid(row=22,column=1)
l21=Label(top,text="Last Class Passed :")
l21.grid(row=23,column=1)
m2=Menubutton(top,text="Select from here ",relief=RAISED)
m2.grid(row=23,column=4)
m2.menu = Menu ( m2, tearoff = 0 )
m2["menu"] = m2.menu
m2.menu.add_checkbutton(label="10th",onvalue=1,offvalue=0,variable=var12)
m2.menu.add_checkbutton(label="12th",onvalue=2,offvalue=0,variable=var12)
m2.menu.add_checkbutton(label="ITI",onvalue=3,offvalue=0,variable=var12)
m2.menu.add_checkbutton(label="DIPLOMA",onvalue=4,offvalue=0,variable=var12)
m2.menu.add_checkbutton(label="GRADUATION",onvalue=5,offvalue=0,variable=var12)
m2.menu.add_checkbutton(label="POST GRADUATION",onvalue=6,offvalue=0,variable=var12)

l22=Label(top,text="Board / University :")
l22.grid(row=23,column=8)
m3=Menubutton(top,text="Select from here ",relief=RAISED)
m3.grid(row=23,column=9)
m3.menu = Menu ( m3, tearoff = 0 )
m3["menu"] = m3.menu
m3.menu.add_checkbutton(label="CBSE",onvalue=1,offvalue=0,variable=var17)
m3.menu.add_checkbutton(label="STATE GOVT",onvalue=2,offvalue=0,variable=var17)
m3.menu.add_checkbutton(label="ICSE",onvalue=3,offvalue=0,variable=var17)
m3.menu.add_checkbutton(label="UNIVERSITY",onvalue=4,offvalue=0,variable=var17)
m3.menu.add_checkbutton(label="UNIVERSITY AFF. COLLEGE",onvalue=5,offvalue=0,variable=var17)

l23=Label(top,text="Marks Obtained :")
l23.grid(row=25,column=1)
e12=Entry(top,bd=1,textvariable=var24)
e12.grid(row=25,column=4)
l24=Label(top,text="Total Marks:")
l24.grid(row=25,column=8)
e13=Entry(top,bd=1,textvariable=var25)
e13.grid(row=25,column=9)
l25=Label(top,text="Last year Percentage :")
l25.grid(row=27,column=1)
e17=Entry(top,state=DISABLED)
e17.grid(row=27,column=4)
l26=Label(top,text=" MARKS  DETAILS : ",pady=10)
l26.grid(row=24,column=1)

l27=Label(top,text=" Age : ")
l27.grid(row=14,column=1)
s4=Spinbox(top,from_=18,to=40,width=4,textvariable=var26)
s4.grid(row=14,column=3)

l28=Label(top,text=" Status : ")
l28.grid(row=14,column=7)
c4=Checkbutton(top,text="MARRIED",onvalue=1,offvalue=0,variable=var27)
c4.grid(row=14,column=8)
c5=Checkbutton(top,text="UNMARRIED",onvalue=1,offvalue=0,variable=var22)
c5.grid(row=14,column=9)

l29=Label(top,text=" CONTACT   DETAILS : ",pady=10)
l29.grid(row=19,column=1)
l27=Label(top,text=" Contact no. : ")
l27.grid(row=20,column=1)
e14=Entry(top,bd=1,textvariable=var28)
e14.grid(row=20,column=4)
l28=Label(top,text=" Another Contact no. : ")
l28.grid(row=20,column=9)
e15=Entry(top,bd=1,textvariable=var29)
e15.grid(row=20,column=10)
l30=Label(top,text=" Email : ")
l30.grid(row=21,column=1)
e16=Entry(top,bd=1,textvariable=var30)
e16.grid(row=21,column=4)

l32=Label(top,text="Category :")
l32.grid(row=15,column=1)
m4=Menubutton(top,text="Select from here ",relief=RAISED)
m4.grid(row=15,column=4)
m4.menu = Menu ( m4, tearoff = 0 )
m4["menu"] = m4.menu
m4.menu.add_checkbutton(label="GENERAL",onvalue=1,offvalue=0,variable=var13)
m4.menu.add_checkbutton(label="BCA",onvalue=2,offvalue=0,variable=var13)
m4.menu.add_checkbutton(label="BCB",onvalue=3,offvalue=0,variable=var13)
m4.menu.add_checkbutton(label="SC / ST",onvalue=4,offvalue=0,variable=var13)

l33=Label(top,text="Stream Applied For :")
l33.grid(row=28,column=1)
m5=Menubutton(top,text="Select from here ",relief=RAISED)
m5.grid(row=28,column=4)
m5.menu = Menu ( m5, tearoff = 0 )
m5["menu"] = m5.menu
m5.menu.add_checkbutton(label="CSE",onvalue=1,offvalue=0,variable=var32)
m5.menu.add_checkbutton(label="ECE",onvalue=2,offvalue=0,variable=var32)
m5.menu.add_checkbutton(label="CIVIL",onvalue=3,offvalue=0,variable=var32)
m5.menu.add_checkbutton(label="BIOTECH",onvalue=4,offvalue=0,variable=var32)
m5.menu.add_checkbutton(label="MECHANICAL",onvalue=5,offvalue=0,variable=var32)


c6=Checkbutton(top,text="I have gone through all my details.I am\n sure that all my details are correct .",onvalue=1,offvalue=0,variable=var31)
c6.grid(row=30,column=7)

excel()

b1=Button(top,text="SAVE & NEXT ",bg="green",width=15,command=button1)
b1.grid(row=32,column=9)
b2=Button(top,text="CANCEL ",bg="red",width=10,command=button2)
b2.grid(row=32,column=10)
b2=Button(top,text="PRINT ",bg="orange",width=10,command=button3)
b2.grid(row=32,column=8)

top.mainloop()
